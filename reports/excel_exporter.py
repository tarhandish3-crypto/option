# reports/excel_exporter.py
# -*- coding: utf-8 -*-

"""
ماژول خروجی اکسل (Excel Exporter Module) - نسخه ارتقایافته سیستم تصمیم‌یار (DSS) - معماری V4

این ماژول وظیفه تبدیل ساختار موقعیت‌های کشف‌شده اسکنر به فایل‌های فرمت‌ شده و استاندارد اکسل را 
بر عهده دارد. در نسخه V4، سیستم درون‌یابی ماتریس P&L و نگاشت گام‌های قیمت کاملاً با رفتار 
ماتریس محاسباتی Payoff هماهنگ شده و از فساد ساختاری شیت‌های فرعی جلوگیری می‌کند.
"""

import os
import logging
import numpy as np
import pandas as pd
from typing import List, Optional, Dict, Any
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from core.models import Opportunity
from config import get_price_steps

logger = logging.getLogger("OptionScanner.Reports.ExcelExporter")


class ExcelExporter:
    """
    صادرکننده پیشرفته گزارش‌های نهایی سیستم تصمیم‌یار (DSS) به همراه فیلترینگ و استایل‌دهی هوشمند
    """

    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # دریافت گام‌های استاندارد درصدی سیستم (به عنوان مثال بازه -50 تا +50)
        self.pct_steps = get_price_steps()
        self._initialize_styles()

    def _initialize_styles(self):
        """قالب‌بندی گرافیکی استاندارد گزارش‌های مالی پروژه"""
        self.header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.body_font = Font(name='Segoe UI', size=10)
        self.body_alignment_right = Alignment(horizontal='right', vertical='center')
        self.body_alignment_center = Alignment(horizontal='center', vertical='center')

        # استایل‌های Conditional Formatting
        self.green_font = Font(color='006100', name='Segoe UI', size=10)
        self.green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        self.red_font = Font(color='9C0006', name='Segoe UI', size=10)
        self.red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        self.gold_font = Font(color='B25E00', bold=True, name='Segoe UI', size=10)
        self.gold_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

        self.gray_font = Font(color='808080', italic=True, name='Segoe UI', size=10)

    def export(self, opportunities: List[Opportunity], filename: Optional[str] = None) -> str:
        """تبدیل و ذخیره‌سازی لیست فرصت‌ها به مسیر خروجی اکسل"""
        if not opportunities:
            logger.warning("No opportunities provided for excel export.")
            return ""

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"opportunities_{timestamp}.xlsx"

        filepath = self.output_dir / filename
        self._export_to_excel(opportunities, str(filepath))
        return str(filepath)

    def _export_to_excel(self, opportunities: List[Opportunity], file_path: str):
        """پردازش لایه داده و ساختارسازی فریم‌ورک خروجی"""
        rows_data = []

        for opp in opportunities:
            metadata = opp.metadata if opp.metadata else {}
            
            # استخراج قیمت مرجع بر اساس استاندارد چندلایه بورس ایران
            S0_stock = getattr(opp, 'S0_stock')
        
            # واکشی بردارهای بازدهی ماتریس سود و زیان (P&L Array Mapping)
            pnl_data = metadata.get('returns_monthly_pct', [])
            if not pnl_data:
                pnl_data = metadata.get('net_returns_closed', [])
            
            # بازسازی لایه گام‌های قیمت بر مبنای هماهنگی با Payoff Core
            price_levels = np.array(metadata.get('price_levels', []))
            
            classification = getattr(opp, 'classification', None)
            cls_market = classification.market_type if classification else 'Neutral'
            cls_profile = classification.investor_profile if classification else 'Balanced'
            cls_risk = classification.risk_level if classification else 'Medium'
            cls_desc = classification.description if classification else ''

            # یکپارچه‌سازی توصیف پوزیشن لگ‌ها
            positions_desc = " | ".join([
                f"{leg.contract.ticker if leg.contract else 'Stock'} ({leg.side.value})"
                for leg in opp.legs
            ]) if getattr(opp, 'legs', None) else metadata.get('positions_desc', 'Custom Leg')

            base_info = {
                "Rank": getattr(opp, 'rank'),
                "Strategy": opp.strategy_name,
                "Positions": positions_desc,
                "DTE": getattr(opp, 'days_to_maturity'),
                "Ticker": opp.underlying_ticker,
                "Market Type": cls_market,
                "Investor Profile": cls_profile,
                "Risk Level": cls_risk,

                "Confidence": metadata.get('confidence'),
                "Conservative Score": opp.profile_scores.conservative if opp.profile_scores else 0.0,
                "Balanced Score": opp.profile_scores.balanced if opp.profile_scores else 0.0,
                "Aggressive Score": opp.profile_scores.aggressive if opp.profile_scores else 0.0,
                "Income Score": opp.profile_scores.income if opp.profile_scores else 0.0,
                "Volatility Score": opp.profile_scores.volatility if opp.profile_scores else 0.0,

                "Dynamic Range": self._format_dynamic_range(metadata.get('dynamic_range', []), S0_stock),
                "Expected Value": metadata.get('expected_value', 0.0),
                "Area Ratio": metadata.get('area_ratio', 0.0),
                "POP %": metadata.get('pop', 0.0),
                "Delta": metadata.get('delta', 0.0),
                "Gamma": metadata.get('gamma', 0.0),
                "Theta": metadata.get('theta', 0.0),
                "Vega": metadata.get('vega', 0.0),
                "Sharpe": metadata.get('sharpe_ratio', 0.0),
                "VaR 95%": metadata.get('var_95', 0.0),
                "Gross Max Profit": getattr(opp, 'max_profit', 0.0),
                "Gross Max Loss": getattr(opp, 'max_loss', 0.0),
                "Net Profit (Closed)": metadata.get('net_profit_if_closed', 0.0),
                "Net Profit (Exercised)": metadata.get('net_profit_if_exercised', 0.0),
                "TSE Commission": metadata.get('total_costs_closed', 0.0),
                "Breakeven": self._format_breakeven(metadata.get('break_even_points', [])),
                "Description": cls_desc,
                "Recommendation": metadata.get('recommended_action', 'بررسی سناریوها'),
                "Liquidity": getattr(opp, 'liquidity_score', 0.0),
                "Score": getattr(opp, 'final_score', 0.0)}

            # بازنویسی مکانیسم نگاشت P&L جهت هماهنگی کامل با گام‌های ثابت محاسباتی (V4)
            if len(price_levels) > 1 and len(pnl_data) == len(price_levels):
                # تبدیل سطوح قیمت مطلق به درصدهای انحراف واقعی جهت تطبیق با گام‌های پیکربندی
                actual_pcts = ((np.array(price_levels) / S0_stock) - 1) * 100
                for idx, pct in enumerate(self.pct_steps):
                    # درون‌یابی ایمن بر اساس نوسانات درصدی واقعی
                    val_interp = np.interp(pct, actual_pcts, pnl_data)
                    base_info[f"{pct}%"] = float(round(val_interp, 4))
            else:
                for idx, pct in enumerate(self.pct_steps):
                    base_info[f"{pct}%"] = float(pnl_data[idx]) if idx < len(pnl_data) else 0.0

            rows_data.append(base_info)

        if not rows_data:
            logger.warning("No formatted rows found to build dataframe.")
            return

        df = pd.DataFrame(rows_data)

        # چینش ستون‌های اصلی طبق استانداردهای حسابداری اوراق مشتقه
        main_cols = [
            "Rank", "Strategy", "Positions", "DTE", "Ticker", "Market Type", "Investor Profile", "Risk Level",
            "Confidence", "Conservative Score", "Balanced Score", "Aggressive Score", "Income Score", "Volatility Score",
            "Dynamic Range", "Expected Value", "Area Ratio", "POP %", "Delta", "Gamma", "Theta", "Vega", "Sharpe", "VaR 95%",
            "Gross Max Profit", "Gross Max Loss", "Net Profit (Closed)", "Net Profit (Exercised)",
            "TSE Commission", "Breakeven", "Description", "Recommendation", "Liquidity", "Score"
        ]

        main_cols = [col for col in main_cols if col in df.columns]
        pct_cols = [f"{pct}%" for pct in self.pct_steps if f"{pct}%" in df.columns]
        df = df[main_cols + pct_cols]

        self._write_to_styled_excel(df, file_path)

    def _format_dynamic_range(self, dynamic_range: List[float], S0_stock: float) -> str:
        if not dynamic_range or S0_stock <= 0:
            return "No Safe Range"
        formatted = []
        for price in dynamic_range:
            pct_deviation = ((price / S0_stock) - 1) * 100
            closest_pct = min(self.pct_steps, key=lambda x: abs(x - pct_deviation))
            formatted.append(f"{closest_pct:+.0f}%")
        unique_pcts = sorted(list(set(formatted)), key=lambda x: float(x.replace('%', '')))
        return " | ".join(unique_pcts) if len(unique_pcts) <= 4 else f"[{unique_pcts[0]} ... {unique_pcts[-1]}]"

    def _format_breakeven(self, breakeven_points: List[float]) -> str:
        if not breakeven_points:
            return "None"
        return ", ".join([f"{p:,.0f}" for p in breakeven_points])

    def _write_to_styled_excel(self, df: pd.DataFrame, file_path: str):
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Scanner_Results', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Scanner_Results']

            # اعمال استایل به هدر اصلی
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment

            numeric_cols = [
                "Confidence", "Conservative Score", "Balanced Score", "Aggressive Score", "Income Score", "Volatility Score",
                "Expected Value", "Area Ratio", "POP %", "Delta", "Gamma", "Theta", "Vega",
                "Sharpe", "VaR 95%", "Gross Max Profit", "Gross Max Loss",
                "Net Profit (Closed)", "Net Profit (Exercised)", "TSE Commission", "Liquidity", "Score"
            ]
            pct_cols = [f"{pct}%" for pct in self.pct_steps if f"{pct}%" in df.columns]
            columns_list = df.columns.tolist()

            # تزریق هوشمند استایل بدنه کدهای مالی
            for row_idx, row in enumerate(df.itertuples(index=False), start=2):
                for col_idx, col_name in enumerate(columns_list, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    val = row[col_idx - 1]

                    if val is None or pd.isna(val):
                        cell.value = "-"
                        cell.font = self.gray_font
                        cell.alignment = self.body_alignment_center
                        continue

                    if col_name in numeric_cols:
                        cell.number_format = '0.0%' if col_name == "Confidence" else '#,##0.00'
                        cell.alignment = self.body_alignment_right
                    elif col_name in pct_cols:
                        cell.number_format = '0.00"%"'
                        cell.alignment = self.body_alignment_right
                    else:
                        cell.alignment = self.body_alignment_center

                    cell.font = self.body_font

            # اعمال قواعد شرطی گرافیکی و فعال‌سازی ابزار فیلتر اکسل
            self._apply_conditional_formatting(worksheet, df)
            self._enable_autofilter(worksheet, df)

            # تنظیم خودکار و دقیق عرض ستون‌ها بر اساس طول کاراکترهای فارسی و انگلیسی
            for col in worksheet.columns:
                max_len = 0
                for cell in col:
                    val = str(cell.value or '')
                    actual_len = sum(2 if ord(c) > 128 else 1 for c in val)
                    if actual_len > max_len:
                        max_len = actual_len
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 13), 50)

            # تولید و الحاق ایمن شیت‌های تصمیم‌یار فرعی مپ‌شده با openpyxl
            self._add_summary_sheet(workbook, df)
            self._add_scores_sheet(workbook, df)

    def _apply_conditional_formatting(self, worksheet, df: pd.DataFrame):
        pct_cols = [f"{pct}%" for pct in self.pct_steps if f"{pct}%" in df.columns]
        target_cols = pct_cols + ["Gross Max Profit", "Expected Value", "Net Profit (Closed)", "Net Profit (Exercised)"]

        for col_name in target_cols:
            if col_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            range_str = f"{col_letter}2:{col_letter}{len(df) + 1}"

            rule_pos = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=False)
            rule_pos.fill = self.green_fill
            rule_pos.font = self.green_font
            worksheet.conditional_formatting.add(range_str, rule_pos)

            rule_neg = CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False)
            rule_neg.fill = self.red_fill
            rule_neg.font = self.red_font
            worksheet.conditional_formatting.add(range_str, rule_neg)

        if 'Rank' in df.columns:
            col_letter = get_column_letter(df.columns.get_loc('Rank') + 1)
            rule_rank = CellIsRule(operator='equal', formula=['1'], stopIfTrue=True)
            rule_rank.font = self.gold_font
            rule_rank.fill = self.gold_fill
            worksheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(df) + 1}", rule_rank)

    def _enable_autofilter(self, worksheet, df: pd.DataFrame):
        if len(df) > 0:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    def _add_summary_sheet(self, workbook: Workbook, df: pd.DataFrame):
        """تولید شیت توزیع آماری و خلاصه مدیریتی"""
        if df.empty:
            return

        summary_ws = workbook.create_sheet(title='Summary')
        summary_ws.append(['شاخص یا سناریوی تحلیلی بازار اوراق اختیار', 'مقدار محاسباتی / فراوانی توزیع سیستم'])
        summary_ws.append(['کل موقعیت‌های معاملاتی تحلیل‌شده', len(df)])

        numeric_cols = [
            ("Expected Value", "میانگین ارزش مورد انتظار (EV)"),
            ("POP %", "میانگین احتمال موفقیت پوزیشن‌ها (POP)"),
            ("Sharpe", "میانگین نسبت شارپ سبد فرصت‌ها"),
            ("Confidence", "میانگین ضریب اطمینان طبقه‌بندی مدل"),
            ("Net Profit (Closed)", "میانگین سود خالص در پوزیشن‌های بسته‌شده"),
        ]

        for col_name, fa_label in numeric_cols:
            if col_name in df.columns:
                # کپی امن و پاکسازی داده جهت جلوگیری از خطای آماری پانداس
                col_series = df[col_name].copy().replace('-', np.nan)
                col_df = pd.to_numeric(col_series, errors='coerce')
                if not col_df.empty and col_df.notna().any():
                    summary_ws.append([fa_label, round(float(col_df.mean()), 2)])

        summary_ws.append(['', ''])
        summary_ws.append(['تفکیک و توزیع موقعیت‌ها بر اساس وضعیت روند بازار', 'تعداد پوزیشن'])

        for col_target in ['Market Type', 'Risk Level', 'Investor Profile']:
            if col_target in df.columns:
                counts = df[col_target].value_counts()
                for category, count in counts.items():
                    summary_ws.append([f"تعداد پوزیشن در کلاس {category}", int(count)])

        # استایل‌دهی شیت خلاصه
        for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.font = self.body_font
                if cell.row in [1, 2 + len(numeric_cols) + 2]:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                else:
                    cell.alignment = self.body_alignment_right if cell.column == 2 else self.body_alignment_center

        summary_ws.column_dimensions['A'].width = 45
        summary_ws.column_dimensions['B'].width = 35

    def _add_scores_sheet(self, workbook: Workbook, df: pd.DataFrame):
        """ساخت شیت ماتریسی امتیازدهی چندگانه پروفایل ریسک"""
        if df.empty:
            return

        scores_ws = workbook.create_sheet(title='Profile_Scores')
        score_cols = [
            "Rank", "Strategy", "Positions", "Ticker", "Confidence",
            "Conservative Score", "Balanced Score", "Aggressive Score", "Income Score", "Volatility Score"
        ]
        score_cols = [col for col in score_cols if col in df.columns]
        sub_df = df[score_cols]

        # ایجاد هدر شیت امتیازها به صورت کاملاً هماهنگ با openpyxl
        for c_idx, col_name in enumerate(score_cols, start=1):
            cell = scores_ws.cell(row=1, column=c_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        # بارگذاری ردیف‌ها
        for r_idx, row in enumerate(sub_df.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = scores_ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = self.body_font

                col_name = score_cols[c_idx - 1]
                if col_name in ["Confidence", "Conservative Score", "Balanced Score", "Aggressive Score", "Income Score", "Volatility Score"]:
                    cell.number_format = '0.0%' if col_name == "Confidence" else '#,##0.00'
                    cell.alignment = self.body_alignment_right
                else:
                    cell.alignment = self.body_alignment_center

        score_fields = ["Conservative Score", "Balanced Score", "Aggressive Score", "Income Score", "Volatility Score"]
        for field in score_fields:
            if field in score_cols:
                c_idx = score_cols.index(field) + 1
                c_letter = get_column_letter(c_idx)
                range_str = f"{c_letter}2:{c_letter}{len(df) + 1}"

                rule_high = CellIsRule(operator='greaterThan', formula=['70.0'], stopIfTrue=False)
                rule_high.fill = self.green_fill
                rule_high.font = self.green_font
                scores_ws.conditional_formatting.add(range_str, rule_high)

        for col in scores_ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                actual_len = sum(2 if ord(c) > 128 else 1 for c in val)
                if actual_len > max_len:
                    max_len = actual_len
            col_letter = get_column_letter(col[0].column)
            scores_ws.column_dimensions[col_letter].width = max(max_len + 4, 15)