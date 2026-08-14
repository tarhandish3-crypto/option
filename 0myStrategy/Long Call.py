# long_call_strategy.py
# -*- coding: utf-8 -*-


import sys
from pathlib import Path

# تنظیم مسیر پروژه
current_file_path = Path(__file__).resolve()
current_dir = current_file_path.parent
root_dir = current_dir.parent
sys.path.append(str(root_dir))

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from datetime import datetime
from data.downloader import MarketDownloader
from data.cleaner import DataCleaner
from config import (
    get_commission_rate,
    get_exercise_fee_rate,
    get_symbol_kind,
    get_symbol_market,)


def long_call_with_fees(premium_call, stock_price, strike_price, contract_size,
                        opt_buy_commission, exercise_fee_rate, days):
    """
    محاسبه بازده استراتژی Long Call با احتساب کارمزدها
    """

    # ========== 1. محاسبه هزینه‌های ورود ==========
    premium_total = -round(premium_call * contract_size, 0)
    entry_fee = round(premium_total * opt_buy_commission, 0)

    # ========== 2. سرمایه اولیه (خروج نقدی) ==========
    # منفی چون پول پرداخت می‌کنیم.
    initial_investment = premium_total + entry_fee

    # ========== 3. محاسبه سود ناخالص در سررسید ==========
    # اگر قیمت پایه بالاتر از قیمت اعمال باشد، سود داریم
    intrinsic_value = max(0, stock_price - strike_price) * contract_size

    # ========== 4. کارمزد اعمال (فقط در صورت سوددهی) ==========
    exercise_fee = 0
    if stock_price > strike_price:
        settlement_amount = strike_price * contract_size
        exercise_fee = -round(settlement_amount * exercise_fee_rate, 0)

    # ========== 5. سود خالص نهایی ==========
    # سود خالص = ارزش ذاتی - حق‌الزام پرداختی - کارمزدها
    net_profit = intrinsic_value + initial_investment + exercise_fee

    # ========== 6. بازده درصدی ==========
    profit_percent = round((net_profit / abs(initial_investment))
                           * 100, 2) if initial_investment != 0 else 0
    monthly_return = round(profit_percent * (30 / days), 2)

    # ========== 7. نقطه سربه‌سر (قیمت پایه در سررسید) ==========
    # باید هزینه کارمزدها را نیز به حق‌الزام اضافه کنیم
    total_cost_per_share = premium_call + (abs(entry_fee) / contract_size) + (
        abs(exercise_fee) / contract_size) if stock_price > strike_price else premium_call
    break_even_price = round(strike_price + total_cost_per_share, 0)

    # ========== 8. درصد فاصله قیمت پایه فعلی تا نقطه سربه‌سر ==========
    # نشان می‌دهد قیمت پایه فعلی چند درصد با نقطه سربه‌سر فاصله دارد
    if break_even_price != 0:
        # اگر break_even_price > stock_price: یعنی باید قیمت بالا برود (درصد مثبت)
        # اگر break_even_price < stock_price: یعنی قیمت پایین‌تر از نقطه سربه‌سر است (درصد منفی)
        break_even_percent = round(
            ((break_even_price - stock_price) / stock_price) * 100, 2)
    else:
        break_even_percent = 0

    return {
        'net_profit': net_profit,
        'profit_percent': profit_percent,
        'monthly_return': monthly_return,
        'break_even_price': break_even_price,
        'break_even_percent': break_even_percent,
        'intrinsic_value': intrinsic_value,
        'fees_total': entry_fee + exercise_fee}


def load_and_filter_data():
    """
    بارگذاری داده‌ها از TSE و فیلتر کردن گزینه‌های اختیار خرید
    """
    # بارگذاری داده‌ها
    df_raw = MarketDownloader.from_tsetmc_direct()
    df_cleaned = DataCleaner.clean(df_raw)
    df_final = DataCleaner.add_derived_columns(df_cleaned)

    # تغيير قيمت نماد پايه برای نمادی که قبلا داخل ‍رتفو دارم و خرید
    # UnderlyingTicker = 'خودرو'
    # if UnderlyingTicker in df_final['UnderlyingTicker'].values:
    #     mask_self = df_final['UnderlyingTicker'] == UnderlyingTicker
    #     df_final.loc[mask_self, 'UnderlyingPrice'] = 528
    # Ticker = 'ضفزر508'
    # df_final = df_final[df_final['Ticker'] == Ticker].reset_index(drop=True)
    # df_final['AskPrice'] = 27600

    # فیلتر کردن گزینه‌های اختیار خرید
    filter_option = df_final[
        (df_final['DaysToMaturity'] > 2.0) &
        (df_final['Type'].apply(lambda x: x.name == 'CALL'))].copy()

    # حذف موارد نامطلوب
    EXCLUDED_UNDERLYING = ['اهرم']
    EXCLUDED_NAME_PATTERN = ['1405/04', '1405-04']
    exclude_mask = (
        (filter_option['UnderlyingTicker'].isin(EXCLUDED_UNDERLYING)) &
        (filter_option['Name'].str.contains(
            '|'.join(EXCLUDED_NAME_PATTERN), na=False)))

    filter_option = filter_option[~exclude_mask].copy()
    # filter_option = filter_option[filter_option['UnderlyingTicker'].isin(EXCLUDED_UNDERLYING)]

    return filter_option


def run_long_call_strategy(df_options, max_break_even_percent=12):
    """
    اجرای استراتژی Long Call روی داده‌های فیلتر شده
    """
    results_fee = []

    for underlying_symbol, group in df_options.groupby('UnderlyingTicker'):
        market = get_symbol_market(underlying_symbol)
        kind = get_symbol_kind(underlying_symbol)

        opt_buy_commission = get_commission_rate(market, 'option', True)
        exercise_fee_rate = get_exercise_fee_rate(market, kind)

        for _, item in group.iterrows():
            # استخراج اطلاعات مورد نیاز
            ticker = item['Ticker']
            strike_price = item['StrikePrice']
            premium_call = item['AskPrice']  # قیمت خرید اختیار
            stock_price = item['UnderlyingPrice']
            contract_size = item['ContractSize']
            days = item['DaysToMaturity']

            # محاسبات با کارمزد
            results = long_call_with_fees(
                premium_call, stock_price, strike_price, contract_size,
                opt_buy_commission, exercise_fee_rate, days)

            # ذخیره نتایج
            results_fee.append({
                'underlying': underlying_symbol,
                'stock_price': stock_price,
                'option_symbol': ticker,
                'strike': strike_price,
                'premium': round(premium_call, 0),
                'stock_price': round(stock_price, 0),
                'net_profit': results['net_profit'],
                'profit_percent': results['profit_percent'],
                'monthly_return_%': results['monthly_return'],
                'break_even_price': results['break_even_price'],
                'break_even_percent': results['break_even_percent'],
                'days_to_maturity': days,
                'volume': int(item.get('Volume', 0))})

    result_df = pd.DataFrame(results_fee)

    # فیلتر بر اساس درصد فاصله تا نقطه سربه‌سر
    result_df_filtered = result_df[result_df['break_even_percent']
                                   <= max_break_even_percent].copy()
    result_df_filtered = result_df_filtered.sort_values(
        ['break_even_percent', 'monthly_return_%'], ascending=[True, True]).reset_index(drop=True)

    return result_df_filtered


def save_results_to_excel(result_df, filename="result_long_call.xlsx"):
    """
    ذخیره نتایج در فایل اکسل با استایل‌بندی حرفه‌ای
    """
    # تنظیمات استایل
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78',
                              end_color='1F4E78', fill_type='solid')
    alignment = Alignment(horizontal='center',
                          vertical='center', wrap_text=True)
    body_font = Font(name='Segoe UI', size=10)
    gray_font = Font(color='808080', italic=True, name='Segoe UI', size=10)

    # اضافه کردن timestamp به نام فایل
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_with_time = f"result_long_call_{timestamp}.xlsx"
    filename_with_time = 'result_long_call.xlsx'
    filepath = Path(__file__).parent / filename_with_time

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='long_call', index=False)
        worksheet = writer.sheets['long_call']

        # اعمال استایل به هدر
        for col_idx in range(1, len(result_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # اعمال استایل به بدنه
        columns_list = result_df.columns.tolist()
        for row_idx, row in enumerate(result_df.itertuples(index=False), start=2):
            for col_idx, col_name in enumerate(columns_list, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                val = row[col_idx - 1]

                if val is None or pd.isna(val):
                    cell.value = "-"
                    cell.font = gray_font
                else:
                    cell.font = body_font
                cell.alignment = alignment

        # ========== هایلایت کردن ماکزیمم و مینیمم ستون‌های خاص ==========
        # ستون‌هایی که باید هایلایت شوند
        highlight_columns = [
            'profit_percent',
            'break_even_percent',
            'monthly_return_%',
        ]
        
        # رنگ‌ها
        max_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # سبز
        min_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')  # قرمز
        
        # پیدا کردن اندیس ستون‌ها
        col_indices = {}
        for col_idx, col_name in enumerate(columns_list, start=1):
            if col_name in highlight_columns:
                col_indices[col_name] = col_idx
        
        # برای هر ستون، ماکزیمم و مینیمم را پیدا کن
        for col_name, col_idx in col_indices.items():
            # استخراج مقادیر عددی از ستون (رد کردن ردیف اول که هدر است)
            values = []
            for row_idx in range(2, len(result_df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and cell.value != "-":
                    try:
                        values.append(float(cell.value))
                    except:
                        pass
            
            if values:
                max_val = max(values)
                min_val = min(values)
                
                # اعمال رنگ به سلول‌های ماکزیمم و مینیمم
                for row_idx in range(2, len(result_df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and cell.value != "-":
                        try:
                            val = float(cell.value)
                            if val == max_val:
                                cell.fill = max_fill
                            elif val == min_val:
                                cell.fill = min_fill
                        except:
                            pass

        # اعمال فیلتر و Freeze Panes
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(result_df.columns))}{len(result_df) + 1}"
        worksheet.freeze_panes = 'A2'

        # ========== تنظیم خودکار عرض ستون‌ها ==========
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                if cell.value:
                    text = str(cell.value)
                    if '\n' in text:
                        lines = text.split('\n')
                        line_length = max(len(line) for line in lines)
                    else:
                        line_length = len(text)

                    if line_length > max_length:
                        max_length = line_length

            adjusted_width = min(max_length + 5, 50)
            worksheet.column_dimensions[column].width = adjusted_width

    print(f"result {filename_with_time}  save")
    return filename_with_time


def main():
    """
    تابع اصلی اجرای استراتژی Long Call
    """
    try:
        # 1. بارگذاری و فیلتر داده‌ها
        filtered_data = load_and_filter_data()

        if filtered_data.empty:
            return

        # 2. اجرای استراتژی
        results = run_long_call_strategy(
            filtered_data, max_break_even_percent=12)

        if results.empty:
            return

        # 3. ذخیره نتایج
        save_results_to_excel(results)

    except Exception as e:
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
