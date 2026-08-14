# Long Call - kala.py
# -*- coding: utf-8 -*-

import warnings
import logging
import requests
import pandas as pd
import re
import jdatetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path

# تنظیم مسیر پروژه
current_file_path = Path(__file__).resolve()
current_dir = current_file_path.parent
root_dir = current_dir.parent
sys.path.append(str(root_dir))


# =====================================================================
# بخش 1: تنظیمات اولیه
# =====================================================================

warnings.filterwarnings('ignore')
logger = logging.getLogger(__name__)


# =====================================================================
# بخش 2: نگاشت‌ها و کارمزدهای بورس کالا
# =====================================================================

# نگاشت نام‌های صندوق‌ها در بورس کالا به نام‌های دقیق در بورس
FUND_MAPPING = {
    'صندوق طلای کهربا': 'صندوق س. كالاي كهربا',
    'صندوق طلای درخشان آبان': 'صندوق س.كالاي آبان',
    'صندوق طلای دنای زاگرس': 'صندوق س.كالاي دناي زاگرس',
    'صندوق طلای کارآمد': 'صندوق س.كالاي كارآمد',
    'صندوق طلای پارسیان': 'صندوق س. كالاي پارسيان',
    'شمش طلا': 'شمش طلا',
    'شمش نقره': 'شمش نقره',
}

# نگاشت نام AssetName به نوع دارایی برای دریافت کارمزد
ASSET_TYPE_MAPPING = {
    'صندوق طلای کهربا': 'صندوق طلا',
    'صندوق طلای درخشان آبان': 'صندوق طلا',
    'صندوق طلای دنای زاگرس': 'صندوق طلا',
    'صندوق طلای کارآمد': 'صندوق طلا',
    'صندوق طلای پارسیان': 'صندوق طلا',
    'شمش طلا': 'شمش طلا',
    'شمش نقره': 'شمش نقره',
}

# =====================================================================
# بخش 3: کارمزدهای بورس کالا
# =====================================================================

# کارمزد معاملات قراردادهای اختیار معامله بورس کالا
# بر اساس ارزش هر قرارداد (قیمت معامله شده × اندازه قرارداد)
IME_COMMISSION = {
    # کارمزد معامله
    'trade_buyer': 0.0012,      # 0.0008 + 0.0004
    'trade_seller': 0.0012,     # 0.0008 + 0.0004
    # کارمزد تسویه و تحویل
    'exercise_buyer': 0.0014,   # 0.0004 + 0.001
    'exercise_seller': 0.0014,  # 0.0004 + 0.001
}

# کارمزد بر اساس نوع دارایی پایه (برای موارد خاص)
IME_COMMISSION_BY_ASSET = {
    'سکه طلا': {
        'trade_buyer': 0.00136,     # 0.0008 + 0.0004 + 0.00016
        'trade_seller': 0.00136,    # 0.0008 + 0.0004 + 0.00016
        'exercise_buyer': 0.0014,
        'exercise_seller': 0.0014,
    },
    'صندوق طلا': {
        'trade_buyer': 0.0012,
        'trade_seller': 0.0012,
        'exercise_buyer': 0.0014,
        'exercise_seller': 0.0014,
    },
    'شمش طلا': {
        'trade_buyer': 0.0012,
        'trade_seller': 0.0012,
        'exercise_buyer': 0.0014,
        'exercise_seller': 0.0014,
    },
    'شمش نقره': {
        'trade_buyer': 0.0012,
        'trade_seller': 0.0012,
        'exercise_buyer': 0.0014,
        'exercise_seller': 0.0014,
    },
    'default': {
        'trade_buyer': 0.0012,
        'trade_seller': 0.0012,
        'exercise_buyer': 0.0014,
        'exercise_seller': 0.0014,
    }
}


def get_ime_commission(asset_type: str, is_buy: bool = True, is_exercise: bool = False):
    """
    دریافت نرخ کارمزد بورس کالا
    """
    side = 'buyer' if is_buy else 'seller'

    if is_exercise:
        key = f'exercise_{side}'
    else:
        key = f'trade_{side}'

    # اگر نوع دارایی در دیکشنری باشد، از آن استفاده کن
    if asset_type in IME_COMMISSION_BY_ASSET:
        return IME_COMMISSION_BY_ASSET[asset_type].get(key, IME_COMMISSION[key])

    # در غیر این صورت از مقدار پیش‌فرض استفاده کن
    return IME_COMMISSION_BY_ASSET['default'].get(key, IME_COMMISSION[key])


# =====================================================================
# بخش 4: توابع استخراج داده از بورس کالا
# =====================================================================


def extract_strike_from_description(description: str) -> float:
    """
    استخراج قیمت اعمال از توضیحات قرارداد
    مثال: '... با قیمت اعمال 22,000 ریال' -> 22000
    """
    if pd.isna(description):
        return 0.0

    desc = str(description)

    # الگوی قیمت اعمال: اعداد با کاما یا بدون کاما
    pattern = r'قیمت اعمال\s*([\d,،]+)'

    match = re.search(pattern, desc)
    if match:
        # حذف تمام کاماها (فارسی و انگلیسی)
        strike_str = match.group(1).replace(',', '').replace('،', '')
        try:
            return float(strike_str)
        except:
            return 0.0

    return 0.0


def extract_fund_name(description: str) -> str:
    """
    استخراج نام دارایی پایه از توضیحات قرارداد
    بر اساس ساختار: قرارداد اختیار معامله [خرید/فروش] [دارایی پایه] سررسید ...
    """
    if pd.isna(description):
        return ""

    desc = str(description)

    # حذف عبارت ابتدایی
    cleaned = re.sub(r'^قرارداد اختیار معامله (خرید|فروش)\s+', '', desc)
    cleaned = re.sub(r'واحدهای سرمایه گذاری\s+', '', cleaned)

    # استخراج تا کلمه "سررسید"
    match = re.search(r'^(.*?)\s+سررسید', cleaned)
    if match:
        return match.group(1).strip()

    return ""


def extract_option_type(contract_code: str) -> str:
    """
    تشخیص نوع اختیار از روی کد قرارداد
    """
    if pd.isna(contract_code):
        return 'Unknown'

    code = str(contract_code)

    if 'C' in code and 'P' not in code:
        return 'Call'
    elif 'P' in code:
        return 'Put'
    else:
        return 'Unknown'


def calculate_days_to_maturity(delivery_date: str, last_trade_date: str) -> int:
    """
    محاسبه تعداد روزهای مانده تا سررسید
    """
    if pd.isna(delivery_date) or pd.isna(last_trade_date):
        return 0

    try:
        # تبدیل تاریخ شمسی به میلادی
        delivery_parts = delivery_date.split('/')
        last_trade_parts = last_trade_date.split('/')

        delivery_gregorian = jdatetime.date(
            int(delivery_parts[0]),
            int(delivery_parts[1]),
            int(delivery_parts[2])).togregorian()

        last_trade_gregorian = jdatetime.date(
            int(last_trade_parts[0]),
            int(last_trade_parts[1]),
            int(last_trade_parts[2])).togregorian()

        # اختلاف به روز
        delta = delivery_gregorian - last_trade_gregorian
        return delta.days

    except:
        return 0


def summarize_contract(group, last_market_date):
    """
    خلاصه‌سازی داده‌های هر قرارداد
    """
    # مرتب‌سازی بر اساس تاریخ (جدیدترین آخر)
    group = group.sort_values('DT_en', ascending=False)

    # پیدا کردن ردیف مربوط به آخرین روز بازار
    latest_day_row = group[group['DT_en'] == last_market_date]

    if not latest_day_row.empty:
        # اگر در آخرین روز بازار معامله داشته، از آن استفاده کن
        latest = latest_day_row.iloc[0].copy()
    else:
        # اگر در آخرین روز بازار معامله نداشته،
        # از آخرین روزی که معامله داشته استفاده کن
        latest = group.iloc[0].copy()

        # اما تاریخ را به آخرین روز بازار تغییر بده
        # و حجم و ارزش را صفر کن (چون در آن روز معامله نداشته)
        latest['TradesVolume'] = 0
        latest['TradesValue'] = 0
        latest['DT_en'] = last_market_date

    # تبدیل آخرین روز به دیکشنری
    summary = latest.to_dict()

    # محاسبه مجموع و میانگین
    total_volume = group['TradesVolume'].sum()
    days_count = len(group)

    # میانگین روزانه
    summary['AvgTradesVolume'] = round(
        total_volume / days_count, 0) if days_count > 0 else 0

    return pd.Series(summary)


def fetch_kala_data(from_date='1405/05/15', to_date='1406/5/22'):
    """
    دریافت داده‌های اختیار معامله از بورس کالا
    """
    print("Fetching data from IME (Bours Kala)...")

    url = "https://www.ime.co.ir/subsystems/ime/option/optionboarddata.ashx"

    params = {
        'f': from_date,
        't': to_date,
        'c': '-1',
        'ot': '0',
        'lang': '8',
        'order': 'asc',
        'offset': '0', }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:153.0) Gecko/20100101 Firefox/153.0', }

    # درخواست GET با استفاده از Session برای حفظ کوکی‌ها
    session = requests.Session()
    response = session.get(url, params=params, headers=headers)

    # بررسی وضعیت درخواست
    response.raise_for_status()

    # تبدیل به JSON
    response_data = response.json()

    rows = response_data.get('rows', [])
    df_rows = pd.DataFrame(rows)

    if df_rows.empty:
        print("No data received from IME.")
        return df_rows

    # افزودن ستون‌های محاسباتی
    df_rows['StrikePrice'] = df_rows['ContractDescription'].apply(
        extract_strike_from_description)
    df_rows['AssetName'] = df_rows['ContractDescription'].apply(
        extract_fund_name)
    df_rows['OptionType'] = df_rows['ContractCode'].apply(extract_option_type)
    df_rows['DaysToMaturity'] = df_rows.apply(
        lambda row: calculate_days_to_maturity(row['DeliveryDate'], row['DT']), axis=1)
    df_rows['DT_en'] = pd.to_datetime(df_rows['DT_en'], errors='coerce')

    # محاسبه آخرین روز معاملاتی کل بازار
    last_market_date = df_rows['DT_en'].max()

    # خلاصه‌سازی هر قرارداد
    df_rows = df_rows.groupby('ContractCode', group_keys=False).apply(
        lambda g: summarize_contract(g, last_market_date), include_groups=False).reset_index()

    # حذف قراردادهای بدون معامله
    df_rows = df_rows[df_rows['AvgTradesVolume'] > 0]

    # مرتب‌سازی
    df_rows = df_rows.sort_values(
        ['AssetName', 'OptionType', 'StrikePrice'],
        ascending=[True, True, True]).reset_index(drop=True)

    return df_rows


# =====================================================================
# بخش 5: توابع دریافت قیمت پایه از بورس
# =====================================================================

def fetch_underlying_prices():
    """
    دریافت قیمت پایه صندوق‌های طلا از بورس
    """
    print("Fetching underlying prices from TSE...")

    url = 'https://old.tsetmc.com/tsev2/data/MarketWatchInit.aspx?h=0&r=0'
    darkhast = requests.get(url, timeout=30)
    main_text = darkhast.text
    csvs = main_text.split('@')
    main_csv = csvs[2]
    csv = main_csv.split(';')
    rows = [row.split(',') for row in csv]

    columns = ['ID', 'ISO', 'nemad', 'name', '4', 'open', 'payani',
               'last', "NumberOfTrades", 'Volume', 'arzesh_trade',
               'baze_min_rooz', 'baze_max_rooz', 'Yesterday_Price',
               'EPS', 'Base_Volume', '16', '17', '18', 'geymat_mojaz_max',
               'geymat_mojaz_min', 'Number_saham', 'group_code', 'extra1', 'extra2', 'extra3']

    df_rows_bourse = pd.DataFrame(rows, columns=columns)
    df_rows_bourse = df_rows_bourse[df_rows_bourse['group_code'] == '380'].reset_index(
        drop=True)

    # فقط نمادهای فعال (QS) را نگه دار
    df_rows_bourse_filtered = df_rows_bourse[df_rows_bourse['extra3'] == 'QS'].copy(
    )

    # تغییر: ایجاد دیکشنری با کلید name و مقدار nemad
    fund_nemad = df_rows_bourse_filtered.set_index('name')['nemad'].to_dict()
    # دیکشنری قیمت‌ها
    fund_prices = df_rows_bourse_filtered.set_index('name')['last'].to_dict()

    return fund_nemad, fund_prices


def add_underlying_prices_to_kala(df_kala, fund_nemad, fund_prices):
    """
    تزریق قیمت پایه به دیتافریم بورس کالا
    """

    def get_underlying_price(asset_name):
        if asset_name in FUND_MAPPING:
            symbol = FUND_MAPPING[asset_name]
            if symbol and symbol in fund_prices:
                try:
                    return float(fund_prices[symbol])
                except:
                    return 0.0
        return 0.0

    def get_underlying_nemad(asset_name):
        if asset_name in FUND_MAPPING:
            symbol = FUND_MAPPING[asset_name]
            if symbol and symbol in fund_nemad:
                return fund_nemad[symbol]
        return asset_name

    df_kala['UnderlyingPrice'] = df_kala['AssetName'].apply(
        get_underlying_price)
    df_kala['UnderlyingNemad'] = df_kala['AssetName'].apply(
        get_underlying_nemad)

    # نمایش آمار قیمت‌های تزریق شده
    price_count = (df_kala['UnderlyingPrice'] > 0).sum()
    print(
        f"Underlying price added for {price_count} out of {len(df_kala)} records")

    return df_kala


# =====================================================================
# بخش 6: تابع محاسبه Long Call برای بورس کالا
# =====================================================================

def long_call_with_fees_kala(premium_call, stock_price, strike_price, contract_size,
                             opt_buy_commission, exercise_fee_rate, days):
    """
    محاسبه بازده استراتژی Long Call برای بورس کالا با احتساب کارمزدها
    """
    # ========== 1. محاسبه هزینه‌های ورود ==========
    premium_total = -round(premium_call * contract_size, 0)
    entry_fee = round(premium_total * opt_buy_commission, 0)

    # ========== 2. سرمایه اولیه ==========
    initial_investment = premium_total + entry_fee

    # ========== 3. محاسبه سود ناخالص در سررسید ==========
    intrinsic_value = max(0, stock_price - strike_price) * contract_size

    # ========== 4. کارمزد اعمال ==========
    exercise_fee = 0
    if stock_price > strike_price:
        settlement_amount = strike_price * contract_size
        exercise_fee = -round(settlement_amount * exercise_fee_rate, 0)

    # ========== 5. سود خالص نهایی ==========
    net_profit = intrinsic_value + initial_investment + exercise_fee

    # ========== 6. بازده درصدی ==========
    profit_percent = round((net_profit / abs(initial_investment))
                           * 100, 2) if initial_investment != 0 else 0
    monthly_return = round(profit_percent * (30 / days), 2)

    # ========== 7. نقطه سربه‌سر ==========
    total_cost_per_share = premium_call + (abs(entry_fee) / contract_size) + (
        abs(exercise_fee) / contract_size) if stock_price > strike_price else premium_call
    break_even_price = round(strike_price + total_cost_per_share, 0)

    # ========== 8. درصد فاصله تا نقطه سربه‌سر ==========
    if break_even_price != 0:
        break_even_percent = round(
            ((break_even_price - stock_price) / stock_price) * 100, 2)
    else:
        break_even_percent = 0

    # ========== 9. بیشترین ضرر ممکن ==========

    # درصد بیشترین ضرر نسبت به قیمت فعلی سهام پایه
    # نشان می‌دهد که قیمت پایه چند درصد باید کاهش یابد تا به نقطه بیشترین ضرر برسد
    # بیشترین ضرر زمانی رخ می‌دهد که قیمت پایه به زیر قیمت اعمال برسد
    if stock_price > 0 and strike_price > 0:
        # فاصله قیمت فعلی تا قیمت اعمال (به درصد)
        distance_to_strike = ((stock_price - strike_price) / stock_price) * 100
        # اگر قیمت فعلی بالاتر از قیمت اعمال باشد، باید چند درصد کاهش یابد تا به قیمت اعمال برسد
        if distance_to_strike > 0:
            max_loss_price_percent = round(distance_to_strike, 2)
        else:
            # اگر قیمت فعلی پایین‌تر از قیمت اعمال باشد، در حال حاضر در منطقه بیشترین ضرر هستیم
            max_loss_price_percent = 0.0
    else:
        max_loss_price_percent = 0.0

    # ========== 10. درصد ریسک نسبت به اختلاف قیمت (حاشیه امنیت) ==========
    if stock_price > strike_price:
        price_difference = stock_price - strike_price
        if price_difference > 0:
            total_premium_cost = premium_call * contract_size + abs(entry_fee)
            risk_percent = round((total_premium_cost / (price_difference * contract_size)) * 100, 2)
        else:
            risk_percent = 100.0
    else:
        risk_percent = 100.

    return {
        'net_profit': net_profit,
        'profit_percent': profit_percent,
        'monthly_return': monthly_return,
        'break_even_price': break_even_price,
        'break_even_percent': break_even_percent,
        'intrinsic_value': intrinsic_value,
        'fees_total': entry_fee + exercise_fee,
        'max_loss_price_percent': max_loss_price_percent,
        'risk_percent': risk_percent,
    }


# =====================================================================
# بخش 7: توابع اصلی اجرای استراتژی
# =====================================================================

def load_and_filter_kala_data(from_date='1405/05/15', to_date='1406/5/22'):
    """
    بارگذاری و فیلتر کردن داده‌های بورس کالا با تزریق قیمت پایه
    """
    # 1. دریافت داده از بورس کالا
    df_kala = fetch_kala_data(from_date, to_date)

    if df_kala.empty:
        print("No data received from IME.")
        return pd.DataFrame()

    # 2. دریافت قیمت پایه از بورس
    fund_nemad, fund_prices = fetch_underlying_prices()

    # 3. تزریق قیمت پایه به دیتافریم بورس کالا
    df_kala = add_underlying_prices_to_kala(df_kala, fund_nemad, fund_prices)

    # 4. فیلتر کردن CALL
    df_call = df_kala[df_kala['OptionType'] == 'Call'].copy()

    # فیلتر بر اساس روز تا سررسید
    df_call = df_call[df_call['DaysToMaturity'] > 2.0].copy()

    # حذف قراردادهای بدون قیمت معتبر
    df_call = df_call[df_call['LastPrice'] > 0].copy()
    df_call = df_call[df_call['StrikePrice'] > 0].copy()
    # فقط قراردادهایی که قیمت پایه دارند
    df_call = df_call[df_call['UnderlyingPrice'] > 0].copy()

    return df_call


def run_long_call_kala_strategy(df_options, max_break_even_percent=25):
    """
    اجرای استراتژی Long Call روی داده‌های بورس کالا
    """
    print("Running Long Call strategy for Kala...")

    results_fee = []

    for _, item in df_options.iterrows():
        # استخراج اطلاعات
        contract_description = item.get('ContractDescription', '')
        ContractCode = item.get('ContractCode', '')
        strike_price = item.get('StrikePrice', 0)
        premium_call = item.get('LastPrice', 0)  # قیمت آخرین معامله
        stock_price = item.get('UnderlyingPrice', 0)
        contract_size = 1  # اندازه قرارداد در بورس کالا معمولاً 1 است
        days = item.get('DaysToMaturity', 0)
        asset_name = item.get('AssetName', '')
        underlying_nemad = item.get('UnderlyingNemad', '')

        if premium_call <= 0 or stock_price <= 0 or days <= 0:
            continue

        # دریافت نام نماد بورس از FUND_MAPPING
        symbol = FUND_MAPPING.get(asset_name, asset_name)
        # ========== دریافت کارمزد بر اساس نوع دارایی ==========
        asset_type = ASSET_TYPE_MAPPING.get(asset_name, 'default')

        # کارمزد خرید اختیار (برای استراتژی Long Call ما خریدار هستیم)
        opt_buy_commission = get_ime_commission(
            asset_type, is_buy=True, is_exercise=False)
        exercise_fee_rate = get_ime_commission(
            asset_type, is_buy=True, is_exercise=True)

        # محاسبات با کارمزد
        results = long_call_with_fees_kala(
            premium_call, stock_price, strike_price, contract_size,
            opt_buy_commission, exercise_fee_rate, days)

        max_loss_price_percent_scale = results['max_loss_price_percent'] * (
            30 / days)**0.5

        break_even_percent_scale = results['break_even_percent'] * (
                    30 / days)**0.5

        # ذخیره نتایج
        results_fee.append({
            'underlying': asset_name,
            'underlying_TSE': underlying_nemad,
            'ContractCode': ContractCode,
            'stock_price': round(stock_price, 0),
            'option_symbol': contract_description,
            'strike': strike_price,
            'premium': round(premium_call, 0),
            'net_profit': results['net_profit'],
            'profit_percent': results['profit_percent'],
            'monthly_return_%': results['monthly_return'],
            'break_even_price': results['break_even_price'],
            'break_even_percent': results['break_even_percent'],
            'break_even_percent_scale' : round(break_even_percent_scale, 2),
            'max_loss_price_percent': results['max_loss_price_percent'],
            'max_loss_price_percent_scale': round(max_loss_price_percent_scale, 2),
            'risk_percent': results['risk_percent'],
            'days_to_maturity': days,
            'volume': int(item.get('AvgTradesVolume', 0)),
        })

    result_df = pd.DataFrame(results_fee)

    # فیلتر بر اساس درصد فاصله تا نقطه سربه‌سر
    if not result_df.empty:
        result_df_filtered = result_df[result_df['break_even_percent_scale']
                                       <= max_break_even_percent].copy()
        result_df_filtered = result_df_filtered.sort_values(
            ['break_even_percent', 'monthly_return_%'],
            ascending=[True, True]).reset_index(drop=True)
    else:
        result_df_filtered = result_df

    return result_df_filtered


def save_kala_results_to_excel(result_df, filename="result_long_call_kala.xlsx"):
    """
    ذخیره نتایج بورس کالا در فایل اکسل
    """
    # تنظیمات استایل
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78',
                              end_color='1F4E78', fill_type='solid')
    alignment = Alignment(horizontal='center',
                          vertical='center', wrap_text=True)
    body_font = Font(name='Segoe UI', size=10)
    gray_font = Font(color='808080', italic=True, name='Segoe UI', size=10)

    # ========== تغییر نام ستون‌ها به فارسی (قبل از ذخیره) ==========
    result_df = result_df.rename(columns={
        'stock_price': 'قیمت نماد پایه',
        'monthly_return_%': 'درصد سود ماهانه',
        'break_even_price': 'قیمت سربه‌سر',
        'break_even_percent': 'درصد فاصله تا نقطه سربه‌سر\n(هرچه کمتر = بهتر)',
        'break_even_percent_scale' : 'مقیاس درصد فاصله تا نقطه سربه‌سر\n(هرچه کمتر = بهتر)(به نسبت 30 روز)',
        'max_loss_price_percent': 'درصد فاصله تا زیان حداکثری\n(هرچه بیشتر = امن‌تر)',
        'max_loss_price_percent_scale': 'مقیاس درصد فاصله تا زیان حداکثری\n(هرچه بیشتر = امن‌تر)(به نسبت 30 روز)',
        'volume': 'حجم معاملات',
        'risk_percent': 'درصد ریسک نسبت به حاشیه امنیت\n(هرچه کمتر = بهتر)'
    })

    # ذخیره در پوشه 0myStrategy
    filepath = Path(__file__).parent / filename

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='long_call_kala', index=False)
        worksheet = writer.sheets['long_call_kala']

        # اعمال استایل به هدر
        for col_idx in range(1, len(result_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # اعمال استایل به بدنه
        columns_list = result_df.columns.tolist()
        for row_idx, row in enumerate(result_df.itertuples(index=False), start=2):
            for col_idx, col_name in enumerate(columns_list, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                val = row[col_idx - 1]

                if val is None or pd.isna(val):
                    cell.value = "-"
                    cell.font = gray_font
                else:
                    cell.font = body_font
                cell.alignment = alignment

        # ========== هایلایت کردن ماکزیمم و مینیمم ستون‌های خاص ==========
        # ستون‌هایی که باید هایلایت شوند
        highlight_columns = [
            'درصد سود ماهانه',
            'مقیاس درصد فاصله تا نقطه سربه‌سر\n(هرچه کمتر = بهتر)(به نسبت 30 روز)',
            'مقیاس درصد فاصله تا زیان حداکثری\n(هرچه بیشتر = امن‌تر)(به نسبت 30 روز)',
            'درصد ریسک نسبت به حاشیه امنیت\n(هرچه کمتر = بهتر)',
        ]
        
        # رنگ‌ها
        max_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # سبز
        min_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')  # قرمز
        
        # پیدا کردن اندیس ستون‌ها
        col_indices = {}
        for col_idx, col_name in enumerate(columns_list, start=1):
            if col_name in highlight_columns:
                col_indices[col_name] = col_idx
        
        # برای هر ستون، ماکزیمم و مینیمم را پیدا کن
        for col_name, col_idx in col_indices.items():
            # استخراج مقادیر عددی از ستون (رد کردن ردیف اول که هدر است)
            values = []
            for row_idx in range(2, len(result_df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and cell.value != "-":
                    try:
                        values.append(float(cell.value))
                    except:
                        pass
            
            if values:
                max_val = max(values)
                min_val = min(values)
                
                # اعمال رنگ به سلول‌های ماکزیمم و مینیمم
                for row_idx in range(2, len(result_df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and cell.value != "-":
                        try:
                            val = float(cell.value)
                            if val == max_val:
                                cell.fill = max_fill
                            elif val == min_val:
                                cell.fill = min_fill
                        except:
                            pass

        # اعمال فیلتر و Freeze Panes
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(result_df.columns))}{len(result_df) + 1}"
        worksheet.freeze_panes = 'A2'

        # ========== تنظیم خودکار عرض ستون‌ها ==========
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                if cell.value:
                    text = str(cell.value)
                    if '\n' in text:
                        lines = text.split('\n')
                        line_length = max(len(line) for line in lines)
                    else:
                        line_length = len(text)

                    if line_length > max_length:
                        max_length = line_length

            adjusted_width = min(max_length + 5, 50)
            worksheet.column_dimensions[column].width = adjusted_width

    print(f"Result saved: {filepath}")
    return str(filepath)


def main():
    """
    تابع اصلی اجرای استراتژی Long Call برای بورس کالا
    """
    try:
        # 1. بارگذاری و فیلتر داده‌ها (با تزریق قیمت پایه)
        filtered_data = load_and_filter_kala_data(
            from_date='1405/05/15',
            to_date='1406/5/22')

        if filtered_data.empty:
            print("No data found.")
            return

        # 2. اجرای استراتژی
        results = run_long_call_kala_strategy(
            filtered_data, max_break_even_percent=12)

        if results.empty:
            print("No results found.")
            return

        # 3. ذخیره نتایج
        save_kala_results_to_excel(
            results,
            filename="result_long_call_kala.xlsx")

    except Exception as e:
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
